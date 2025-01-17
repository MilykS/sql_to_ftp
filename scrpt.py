import mysql.connector
from mysql.connector import Error
from openpyxl import load_workbook
from ftplib import FTP
import json



# подключаемся к серверу sql и дб последовательно
def create_connection(host_name, user_name, user_password, db_name):
    connection = None
    try:
        connection = mysql.connector.connect(
            host=host_name,
            user=user_name,
            passwd=user_password,
        )
        print("Connection to SQL server successful")
        try:
            connection = mysql.connector.connect(
                host=host_name,
                user=user_name,
                passwd=user_password,
                database=db_name,
            )
            print("Connection to SQL DB successful")
        except Error as e:
            print(f"The error '{e}' occurred")
    except Error as e:
        print(f"The error '{e}' occurred")
    print(f'TABLES FROM db {db_name}')
    execute_query_show(connection, f'SHOW TABLES FROM {db_name};')
    return connection


# введение запроса к дб
def execute_query(connection, host_name, user_name, user_password):
    flag = 0
    while (flag!=1):
        if (flag==2):
            connection=create_connection(host_name, user_name, user_password, input('cin data base name '))
        query = input(f"cin your query to db '{connection.database}' ")
        cursor = connection.cursor()
        if 'SHOW' in query or 'show' in query:
            execute_query_show(connection, query)
        elif 'SELECT' in query or 'select' in query:
            execute_query_select(connection, query)
        else:
            try:
                cursor.execute(query)
                connection.commit()
                print("Query executed successfully")
            except Error as e:
                print(f"The error '{e}' occurred")
        flag = int(input('is work done?(NO - 0, YES - 1, 2 - CHANGE DB)'))


# введение запроса типа SHOW к дб
def execute_query_show(connection, query):
    cursor = connection.cursor()
    try:
        cursor.execute(query)
        result = cursor.fetchall()
        print(result)
    except Error as e:
        print(f"The error '{e}' occurred")


# запись запроса в эксель файл
def select_to_excel(query_result, querys_columns):
    fn = input('write path to your excel file ')
    # 'C:\\for_ftp\\ex.xlsx'
    wb = load_workbook(fn)
    ws = wb[input('write name of excel list ')]
    # 'data'
    for i in range(len(querys_columns)):
        cell = ws.cell(row=1, column=i + 1)
        cell.value = querys_columns[i]
    wb.save(fn)
    for row in range(len(query_result)):
        for col in range(len(querys_columns)):
            value = query_result[row][col]
            cell = ws.cell(row=row + 2, column=col + 1)
            cell.value = value
    wb.save(fn)
    wb.close()


# Запись запроса в json файл, реализация именно такая в связи с работой метода, исполняющего sql запрос...
# ...(он возвращает только значения получившейся таблички без названий столбцов):...
# ...а без названий столбцов не получится внятной ,,словарной,, структуры json файла.
def select_to_json(query, query_result, querys_columns):
    strings_in_json = []
    for i in range(len(query_result)):# заполнение словаря для JSON файла, где ключи-это названия столбцов(querys_coloumns),
                                    # а значения по ключам-результат запроса(querys_result)
        new = {}
        for j in range(len(querys_columns)):
            new[querys_columns[j]] = query_result[i][j]
        strings_in_json.append(new)
    pre_json = {query: strings_in_json}
    print(pre_json)
    file = open(input('write path to your json file with file extension '), 'w')
    json.dump(pre_json, file, indent=3)
    file.close()

# запись запроса в остальные файлы(.txt и тд)
def select_to_file(query_result, querys_columns):
    file = open(input('write path to your file with file extension '), 'w')
    columns = ''
    for el in querys_columns:
        columns += str(el)
        columns += ' '
    file.write(columns+'\n')
    for i in range(len(query_result)):
        file.write(str(query_result[i])+'\n')
    file.close()


# введение запроса типа SELECT к дб+запись результата запроса в эксель, json или другой вид файла
def execute_query_select(connection, query):
    cursor = connection.cursor()
    try:
        cursor.execute(query)
        result = cursor.fetchall()
        column_names = [description[0] for description in cursor.description]
        print(column_names)
        for i in range(len(result)):
            print(result[i])
        while(True):
            flag = int(input('write it in file?(NO - 0, '
                             'WRITE IN EXCEL FILE - 1, '
                             'WRITE IN OTHER TYPE OF FILE - 2,'
                             'WRITE IN JSON FILE - 3) '))
            if flag == 0:
                return 0
            elif flag == 1:
                select_to_excel(result, column_names)
            elif flag == 2:
                select_to_file(result, column_names)
            elif flag == 3:
                select_to_json(query, result, column_names)
            else:
                pass
    except Error as e:
        print(f"The error '{e}' occurred")


# подключение к фтп, работа с JSON файлами возможна
def FTP_connect(ftp_server, user, password):
    ftp = FTP(ftp_server)
    ftp.login(user=user, passwd=password)
    print(ftp.login())
    ftp.encoding = "utf-8"
    if ftp.passiveserver:
        print('passive mode of data transfers ')
    else:
        print('active mode of data transfers ')
    show_ftp_files(ftp)
    return ftp


# просмотр содержимого на фтп сервере
def show_ftp_files(ftp):
    data = ftp.retrlines('LIST')
    print(data)
    while (True):
        flag = int(input('need browse dirs?(NO - 0, YES - 1, '
                         'CREATE DIR - 2, DELETE DIR - 3) '))
        if flag==1:
            ftp.cwd(input('chose dir '))
            data = ftp.retrlines('LIST')
            print(data)
        elif flag==2:
            ftp.mkd(input('write name of new dir '))
        elif flag==3:
            ftp.rmd(input('write dir name to delete '))
        elif flag==0:
            break
        else:
            pass


# загрузка файла на фтп сервер
def FTP_store(from_file, to_file):
    with open(from_file, 'rb') as file:
        ftp.storbinary(f'STOR {to_file}', file)

    # with open(input('Write path to file for upload '), 'rb') as file:
    #     ftp.storbinary('STOR ','Write name of uploaded file with file extension ', file)


# скачивание файла с фтп сервера
def FTP_retr(from_server, to_pc):
    file = open(to_pc, 'wb')
    ftp.retrbinary(f'RETR {from_server}', file.write)
    file.close()


# удаление файла с фтп сервера
def FTP_delete(file_name):
    ftp.delete(file_name)


# переименование файла на фтп сервере
def FTP_rename_a_file(fromname, toname):
    ftp.rename(fromname, toname)


# исполнение запроса к фтп серверу
def FTP_query(ftp):
    while (True):
        flag = int(input('need work with ftp server? '
                         '(0-to dirs, 1-upload, 2-download, '
                         '3-delete, 4-rename a file, 5-quit) '))
        if flag == 0:
            show_ftp_files(ftp)
        elif flag == 1:
            FTP_store(input('Write path to file for upload with file extension '),
                    input('Write uploaded file name with file extension '))
        elif flag == 2:
            FTP_retr(input('Write file name with file extension for download '),
                     input('Write path to file for download with file extension '))
        elif flag == 3:
            FTP_delete(input('Write file name to delete with file extension '))
        elif flag == 4:
            FTP_rename_a_file(input('Write previous name of file with file extension '),
                              input('Write new name of file with file extension '))
        elif flag == 5:
            ftp.quit()
            break
        else:
            pass


##################################################################################################
##################################################################################################
FLAG=input('quit-0, work with SQL-1, work with FTP server-2 ')
while(FLAG!='0'):

    if (FLAG=='1'):

        host_name = input('cin host_name sql server ')

        user_name = input('cin user_name ')

        user_password = input('cin user_password ')

        db_name = input('cin data base name ')

        connection = create_connection(host_name, user_name, user_password, db_name)

        execute_query(connection, host_name, user_name, user_password)

    elif (FLAG=='2'):

        ftp = FTP_connect(ftp_server=input('cin your FTP server '),
                            user=input('cin user '),
                            password=input('cin password '))

        FTP_query(ftp)

    else:

        pass
    FLAG = input('quit-0, work with SQL-1, work with FTP server-2 ')
